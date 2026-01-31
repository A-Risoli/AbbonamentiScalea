"""Generate test Excel file with 800 rows for import testing."""

import random
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet


def generate_test_excel(output_path: str | Path, num_rows: int = 800):
    """
    Generate test Excel file with sample subscription data.
    
    Follows the new import rules:
    - Start date only on some rows (others blank to test carry-forward)
    - End date can be blank (will auto-set to Dec 31)
    - POS and Bollettino are separate columns with X markers
    
    Args:
        output_path: Path where to save the Excel file
        num_rows: Number of data rows to generate (default 800)
    """
    # Create workbook
    wb = openpyxl.Workbook()
    ws: Worksheet = wb.active
    ws.title = "Abbonamenti"
    
    # Define headers
    headers = [
        "Nome Proprietario",
        "Targa",
        "Data Inizio",
        "Data Fine",
        "Importo",
        "POS",
        "Bollettino",
        "Email",
        "Indirizzo",
        "Cellulare",
    ]
    
    # Write headers with bold font
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
    
    # Sample data for randomization
    first_names = [
        "Mario", "Luigi", "Giuseppe", "Francesco", "Antonio",
        "Giovanni", "Paolo", "Marco", "Luca", "Andrea",
        "Alessandro", "Stefano", "Carlo", "Roberto", "Davide",
        "Maria", "Anna", "Giulia", "Laura", "Sara",
        "Elena", "Chiara", "Francesca", "Rosa", "Angela",
    ]
    
    last_names = [
        "Rossi", "Russo", "Ferrari", "Esposito", "Bianchi",
        "Romano", "Colombo", "Ricci", "Marino", "Greco",
        "Bruno", "Gallo", "Conti", "De Luca", "Costa",
        "Giordano", "Mancini", "Rizzo", "Lombardi", "Moretti",
    ]
    
    streets = [
        "Via Roma", "Via Garibaldi", "Via Mazzini", "Via Cavour",
        "Via Verdi", "Corso Italia", "Piazza Venezia", "Via Dante",
        "Via Firenze", "Corso Vittorio Emanuele", "Via Milano",
        "Via Torino", "Via Napoli", "Piazza del Popolo", "Via XX Settembre",
    ]
    
    # Start date configuration - create clusters of rows with same date
    base_date = datetime(2024, 1, 15)
    current_start_date = base_date
    rows_since_date_change = 0
    
    # Generate data rows
    row_num = 2
    for i in range(num_rows):
        # Every 5-15 rows, add a new start date (simulating grouped entries by day)
        if rows_since_date_change >= random.randint(5, 15) or i == 0:
            # Add new date (advance by some days)
            days_forward = random.randint(1, 30)
            current_start_date = current_start_date + timedelta(days=days_forward)
            rows_since_date_change = 0
            start_date_value = current_start_date.strftime("%d/%m/%Y")
        else:
            # Leave date blank (will carry forward)
            start_date_value = ""
            rows_since_date_change += 1
        
        # Generate owner name
        first_name = random.choice(first_names)
        last_name = random.choice(last_names)
        owner_name = f"{first_name} {last_name}"
        
        # Generate license plate (Italian format)
        plate_letters_1 = "".join(random.choices("ABCDEFGHIJKLMNOPQRSTUVWXYZ", k=2))
        plate_numbers = "".join(random.choices("0123456789", k=3))
        plate_letters_2 = "".join(random.choices("ABCDEFGHIJKLMNOPQRSTUVWXYZ", k=2))
        license_plate = f"{plate_letters_1}{plate_numbers}{plate_letters_2}"
        
        # End date - leave blank 70% of time (will auto-set to Dec 31)
        if random.random() < 0.3:
            # Explicit end date
            end_date_value = datetime(current_start_date.year, 12, 31).strftime("%d/%m/%Y")
        else:
            end_date_value = ""
        
        # Payment amount
        payment_amount = random.choice([50.00, 75.00, 100.00, 120.00, 150.00])
        
        # Payment method - POS or Bollettino (mutually exclusive)
        if random.random() < 0.6:
            # POS
            pos_value = "X"
            bollettino_value = ""
        else:
            # Bollettino
            pos_value = ""
            bollettino_value = "X"
        
        # Optional fields (50% probability)
        if random.random() < 0.5:
            email = f"{first_name.lower()}.{last_name.lower()}@example.com"
        else:
            email = ""
        
        if random.random() < 0.5:
            street = random.choice(streets)
            civic = random.randint(1, 200)
            address = f"{street}, {civic}"
        else:
            address = ""
        
        if random.random() < 0.5:
            mobile = f"3{random.randint(10, 99)}{random.randint(1000000, 9999999)}"
        else:
            mobile = ""
        
        # Write row
        row_data = [
            owner_name,
            license_plate,
            start_date_value,
            end_date_value,
            payment_amount,
            pos_value,
            bollettino_value,
            email,
            address,
            mobile,
        ]
        
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_num, column=col_idx, value=value)
        
        row_num += 1
    
    # Adjust column widths
    ws.column_dimensions["A"].width = 25  # Nome Proprietario
    ws.column_dimensions["B"].width = 12  # Targa
    ws.column_dimensions["C"].width = 15  # Data Inizio
    ws.column_dimensions["D"].width = 15  # Data Fine
    ws.column_dimensions["E"].width = 12  # Importo
    ws.column_dimensions["F"].width = 8   # POS
    ws.column_dimensions["G"].width = 12  # Bollettino
    ws.column_dimensions["H"].width = 30  # Email
    ws.column_dimensions["I"].width = 35  # Indirizzo
    ws.column_dimensions["J"].width = 15  # Cellulare
    
    # Save workbook
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    
    print(f"✓ Generated test file: {output_path}")
    print(f"  - Total rows: {num_rows}")
    print(f"  - Columns: {len(headers)}")
    print(f"  - Date carry-forward: Yes (dates only on first row of each group)")
    print(f"  - End date: Mostly blank (auto-set to Dec 31)")
    print(f"  - Payment: Separate POS/Bollettino columns")


if __name__ == "__main__":
    # Generate test file in assets folder
    output_file = Path(__file__).parent.parent / "assets" / "test1.xlsx"
    generate_test_excel(output_file, num_rows=800)
